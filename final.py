import csv
import openpyxl
from openpyxl.chart import PieChart,Reference
import matplotlib.pyplot as plt

print("layjen6644")

# Assigning variable to .csv file path
path="C:\\FinalExam\\final.csv"

# Function to get user to enter 5 numbers
def askUser():
    # Initializing variable to zero
    total = 0
    # Loop to get 5 numbers
    for i in range(5):
        number = int(input("Enter a whole number: "))
        total += number
    print(f'The total is: {total}')

# Function to get user to enter 5 names and 5 incomes
def askIncome():
    # Loop to get 5 names and 5 incomes
    for i in range(5):
        name = input("Enter a name: ")
        income = input("Enter an income: ")
        # Writing to existing .csv file
        with open(path, "a") as file:
            file.write(name + "," + str(income) + "\n")

# Function to create pie chart
def excelPie():
    # Creating a excel file/workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Reading .csv file
    with open(path, "r") as file:
        reader = csv.reader(file)

        # Retriving data and putting into an .xlsx file
        for row in reader:
            ws.append(row)

    # Finding max rows and columns
    rows = ws.max_row
    cols=ws.max_column
    alphas=[]
    # Creating column letters
    for charCode in range(65,91):
        alphas.append(chr(charCode))
    
    # Creating cell references
    for row in range(1,rows+1):
        for column in range(cols):
            print(ws[alphas[column]+str(row)].value,end="\t")
        print("\n")
        
        
    # Typecasing strings to integers
    for row in range(1,rows+1):
        ws[alphas[1] + str(row)] = int(ws[alphas[1] + str(row)].value)

    # The following will create a pie chart
    # Creates chart object
    myPieChart = PieChart()
    # Title
    myPieChart.title="layjen6644 August 28th, 2026"
    # Getting data and labels
    myData = Reference(ws,min_col=2,min_row=1,max_row=rows)
    myLabels = Reference(ws,min_col=1,min_row=1,max_row=rows)
    # Adding data and labels to chart
    myPieChart.add_data(myData)
    myPieChart.set_categories(myLabels)

    # Adding chart to .xlsx file
    ws.add_chart(myPieChart,"D3")

    # Saving as new .xlsx file
    wb.save("C:\\FinalExam\\final.xlsx")

# Function to create bar graph
def verticalBar():
    names = []
    incomes = []

    # Reading csv file
    with open(path, "r") as file:
        reader = csv.reader(file)

        # Retrieving data from .csv file and putting it in the list variables "names" and "incomes"
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))

    # Creating bar graph
    plt.bar(names, incomes)
    plt.title("layjen6644 August 28th, 2026")
    plt.xlabel("Names")
    plt.ylabel("Income")
    plt.show()
    plt.xticks(rotation=45)

# Calling functions to run
askUser()
askIncome()
excelPie()
verticalBar()
